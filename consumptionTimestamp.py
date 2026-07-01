import pandas as pd
from openpyxl import load_workbook
import os
import glob

# Directory containing the GPI data files
DATA_DIR = r"C:\Users\gokul\Downloads\Final (5) 14\Final (5)\Data\GPI"

SKIPROWS = 5


def write_back_excel(filepath, df, original_header_rows=5):
    """
    Overwrite the data portion of an Excel file (below the original header rows)
    while preserving the top header rows exactly as-is.
    """
    wb = load_workbook(filepath)
    ws = wb.active

    # Remove all rows below the original header
    if ws.max_row > original_header_rows:
        ws.delete_rows(original_header_rows + 1, ws.max_row - original_header_rows)

    # Write new column headers in the next row
    header_row = original_header_rows + 1
    col_names = list(df.columns)
    for col_idx, col_name in enumerate(col_names, 1):
        ws.cell(row=header_row, column=col_idx, value=col_name)

    # Find which column index is "Date" (1-based)
    date_col_idx = col_names.index("Date") + 1 if "Date" in col_names else None

    # Write data rows
    for row_offset, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell_value = value
            # Convert NaT / nan to None so Excel stores blank
            if pd.isnull(value) if not isinstance(value, str) else False:
                cell_value = None
            cell = ws.cell(row=header_row + row_offset - 1, column=col_idx, value=cell_value)
            # Apply Short Date format to the Date column
            if col_idx == date_col_idx and cell_value is not None:
                cell.number_format = "DD-MM-YYYY"

    wb.save(filepath)


def process_consumption_file(consumption_path, booking_path):
    print(f"\nProcessing : {os.path.basename(consumption_path)}")
    print(f"Against    : {os.path.basename(booking_path)}")

    # ── Read files ────────────────────────────────────────────────────────────
    df_con = pd.read_excel(consumption_path, skiprows=SKIPROWS)
    df_book = pd.read_excel(booking_path, skiprows=SKIPROWS)

    # ── Rename Unnamed: 22 ────────────────────────────────────────────────────
    if "Unnamed: 22" in df_con.columns:
        df_con.rename(columns={"Unnamed: 22": "Total Liquid Metal Poured (MT)"}, inplace=True)
        print("  Renamed 'Unnamed: 22' -> 'Total Liquid Metal Poured (MT)'")

    # ── Normalise keys for matching ───────────────────────────────────────────
    for df in [df_con, df_book]:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.normalize()

    df_con["Shift"] = df_con["Shift"].astype(str).str.strip().str.upper()
    df_book["Shift"] = df_book["Shift"].astype(str).str.strip().str.upper()

    # Consumption uses 'Component ID', Booking uses 'ComponentId'
    df_con["Component ID"] = df_con["Component ID"].astype(str).str.strip()
    df_book["ComponentId"] = df_book["ComponentId"].astype(str).str.strip()

    # ── Prepare booking lookup ────────────────────────────────────────────────
    # Sort booking by Date, Shift, ComponentId, StartTime.
    # Detect continuity: consecutive rows for the same component are continuous
    # if EndTime[i] == StartTime[i+1]. A gap means a new separate window.
    # Aggregate each continuous run → first StartTime, last EndTime.
    # Each resulting window gets a rank (0, 1, 2 …) within Date+Shift+ComponentId.

    df_bs = df_book.sort_values(
        ["Date", "Shift", "ComponentId", "StartTime"]
    ).reset_index(drop=True)

    # Identify group boundaries (different Date/Shift/ComponentId)
    same_grp = (
        (df_bs["Date"]        == df_bs["Date"].shift(1)) &
        (df_bs["Shift"]       == df_bs["Shift"].shift(1)) &
        (df_bs["ComponentId"] == df_bs["ComponentId"].shift(1))
    )
    # Continuous = same group AND prev EndTime == curr StartTime (as strings)
    is_continuous = (
        df_bs["EndTime"].shift(1).astype(str).str.strip()
        == df_bs["StartTime"].astype(str).str.strip()
    )
    # A new window starts whenever the group changes OR there is a gap
    df_bs["_win_global"] = (~(same_grp & is_continuous)).cumsum()

    # Aggregate: first StartTime and last EndTime per window
    df_book_agg = (
        df_bs.groupby("_win_global", sort=False)
        .agg(
            Date        =("Date",        "first"),
            Shift       =("Shift",       "first"),
            ComponentId =("ComponentId", "first"),
            StartTime   =("StartTime",   "first"),
            EndTime     =("EndTime",     "last"),
        )
        .reset_index(drop=True)
        .rename(columns={"ComponentId": "Component ID"})
    )

    # Rank windows 0-based per component, ordered by StartTime
    df_book_agg = df_book_agg.sort_values(
        ["Date", "Shift", "Component ID", "StartTime"]
    )
    df_book_agg["_rank"] = df_book_agg.groupby(
        ["Date", "Shift", "Component ID"]
    ).cumcount()

    print(f"  Booking windows after continuity aggregation : {len(df_book_agg)}")

    # Drop any pre-existing StartTime / EndTime in consumption (both naming styles)
    for col in ["StartTime", "EndTime", "Start Time", "End Time"]:
        if col in df_con.columns:
            df_con.drop(columns=[col], inplace=True)

    # ── Rank consumption rows per Date+Shift+Component ID ────────────────────
    # Use Time column if present so row order matches booking window order.
    if "Time" in df_con.columns:
        df_con = df_con.sort_values(
            ["Date", "Shift", "Component ID", "Time"], kind="stable"
        )
    df_con["_rank"] = df_con.groupby(
        ["Date", "Shift", "Component ID"]
    ).cumcount()

    # ── Merge on Date + Shift + Component ID + window rank ───────────────────
    # Continuous runs → merged into 1 window; gapped runs → separate windows.
    # Each consumption occurrence (by rank) maps to the corresponding window.
    df_merged = df_con.merge(
        df_book_agg[["Date", "Shift", "Component ID", "_rank", "StartTime", "EndTime"]],
        on=["Date", "Shift", "Component ID", "_rank"],
        how="left",
    )

    # Fallback: if a consumption row has no matching window (rank exceeds
    # available windows, e.g. 2 consumption rows but only 1 continuous window),
    # assign the last available window for that component.
    unmatched_mask = df_merged["StartTime"].isna()  # still "StartTime" before rename
    if unmatched_mask.any():
        # Build a lookup of the last window per component
        last_window = (
            df_book_agg.sort_values("_rank")
            .groupby(["Date", "Shift", "Component ID"], sort=False)
            .last()
            .reset_index()[["Date", "Shift", "Component ID", "StartTime", "EndTime"]]
            .rename(columns={"StartTime": "_ST_fb", "EndTime": "_ET_fb"})
        )
        df_merged = df_merged.merge(last_window, on=["Date", "Shift", "Component ID"], how="left")
        df_merged.loc[unmatched_mask, "StartTime"] = df_merged.loc[unmatched_mask, "_ST_fb"]
        df_merged.loc[unmatched_mask, "EndTime"]   = df_merged.loc[unmatched_mask, "_ET_fb"]
        df_merged.drop(columns=["_ST_fb", "_ET_fb"], inplace=True)

    df_merged.drop(columns=["_rank"], inplace=True)

    # Rename to spaced column headers for Excel output
    df_merged.rename(columns={"StartTime": "Start Time", "EndTime": "End Time"}, inplace=True)

    matched = df_merged["Start Time"].notna().sum()
    print(f"  Rows matched with StartTime/EndTime : {matched} / {len(df_merged)}")

    write_back_excel(consumption_path, df_merged, original_header_rows=SKIPROWS)
    print(f"  Saved  : {consumption_path}")


def main():
    # Find all Consumption files (exclude ConsumptionBooking files)
    all_xlsx = glob.glob(os.path.join(DATA_DIR, "Consumption_*.xlsx"))
    consumption_files = [f for f in all_xlsx if "booking" not in os.path.basename(f).lower()]

    if not consumption_files:
        print(f"No Consumption_*.xlsx files found in:\n  {DATA_DIR}")
        return

    for consumption_path in sorted(consumption_files):
        basename = os.path.basename(consumption_path)

        booking_basename = basename.replace("Consumption_", "Consumptionbooking_", 1)
        booking_path = os.path.join(DATA_DIR, booking_basename)

        if not os.path.exists(booking_path):
            print(f"\nSkipping {basename}: matching booking file not found.")
            print(f"  Expected: {booking_basename}")
            continue

        process_consumption_file(consumption_path, booking_path)

    print("\nDone.")


if __name__ == "__main__":
    main()
