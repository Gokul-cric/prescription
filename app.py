import os
import json
import glob
import warnings
import smtplib
from datetime import datetime, time, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from typing import List

import pandas as pd
from flask import Flask, request, render_template_string, jsonify
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as col2idx

from preprocess import load_all_configs
from API.helper import additives_simulation
from consumptionTimestamp import process_consumption_file
from preparedsandpreprocess import keep_last_value_only

warnings.filterwarnings("ignore")

# ── Config ────────────────────────────────────────────────────────────────────
configs   = load_all_configs()
config1   = configs.get("config.json")
config2   = configs.get(config1["foundry_name"] + ".json")

cwd          = os.getcwd()
data_dir     = os.path.join(cwd, "Data",          config1["foundry_name"])
results_dir  = os.path.join(cwd, "Model Results", config1["foundry_name"])
analysis_dir = os.path.join(cwd, "Analysis",      config1["foundry_name"])
master_path  = os.path.join(analysis_dir, "Final.xlsx")

PRESCRIPTION_TIMES = [time(0, 0), time(8, 0), time(16, 0)]

SETTINGS_PATH = os.path.join(cwd, "settings.json")

def load_settings():
    with open(SETTINGS_PATH) as f:
        return json.load(f)

def save_settings(data):
    with open(SETTINGS_PATH, "w") as f:
        json.dump(data, f, indent=4)

s        = load_settings()
n        = s["n"]
perShift = s["perShift"]
OFFSET   = s["OFFSET"]
newsand  = s["newsand"]

with open(os.path.join(results_dir, "output2_props.json")) as f:
    model_info = json.load(f)

SUM_COLS = [
    "Core Sand (MT)",
    "Total Prepared Sand (MT)",
    "Total Liquid Metal Poured (MT)",
    "Core Influx Sand (MT)",
]

# ── Final Prediction formula (replaces win32com) ──────────────────────────────
def compute_final_prediction(last_row_data):
    """Compute Final Prediction in pure Python using the formula in Final.xlsx."""
    wb = load_workbook(master_path, data_only=True)
    ws = wb.active

    targets = [ws.cell(1, c).value or 0 for c in range(3, 15)]   # C1:N1
    coeffs  = [ws.cell(2, c).value or 1 for c in range(3, 15)]   # C2:N2
    weights = [ws.cell(3, c).value or 0 for c in range(3, 15)]   # C3:N3
    R1  = ws.cell(1, 18).value or 0    # CORE mean
    R2  = ws.cell(2, 18).value or 1    # CORE coeff
    V1  = ws.cell(1, 22).value or 0    # SMR mean
    V2  = ws.cell(2, 22).value or 1    # SMR coeff
    AC1 = ws.cell(1, 29).value or 1    # new sand coefficient
    wb.close()

    PROPS = [
        "activeClay", "compactibility", "gcs", "gfnAfs",
        "inertFines", "loi", "moisture", "permeability",
        "shearStrength", "splitStrength", "tempOfSandAfterMix", "volatileMatter",
    ]

    # AM:AX — weighted property deviations
    am_ax = [
        (weights[i] * (targets[i] - (last_row_data.get(PROPS[i]) or 0))) / (coeffs[i] or 1)
        for i in range(12)
    ]

    csi = last_row_data.get("csi") or 0
    smr = last_row_data.get("sand_metal_ratio") or 0

    AY = (((csi * perShift / n) - R1) / R2) if R2 else 0   # AY — CSI normalised
    AZ = ((smr - V1) / V2) if V2 else 0                     # AZ — SMR normalised

    am_az_sum = sum(am_ax) + AY + AZ

    bentonite  = last_row_data.get("bentonite_predicted_kg")  or 0
    bentonite1 = last_row_data.get("bentonite_predicted_kg1") or 0
    fss        = 0   # always 0

    AO = am_ax[2]   # gcs deviation
    AP = am_ax[3]   # gfnAfs deviation
    AV = am_ax[9]   # splitStrength deviation

    P1 = bentonite  + am_az_sum + OFFSET + fss / AC1
    P2 = bentonite1 + am_az_sum + OFFSET + fss / AC1 - AO - AP - AV

    return round((P1 + P2) / 2, 2)


# ── Prescription pipeline ─────────────────────────────────────────────────────
def get_prescription_slots(df):
    if df.empty:
        return []
    min_date = df["Start Time stamp"].min().date()
    max_date = df["End Time stamp"].max().date()
    slots, current = [], min_date
    while current <= max_date:
        for t in PRESCRIPTION_TIMES:
            slots.append(datetime.combine(current, t))
        current += timedelta(days=1)
    return sorted(slots)


def build_prescription_records(merged, ps_data):
    merged_sorted = merged.sort_values("End Time stamp").reset_index(drop=True)
    ps_sorted     = ps_data.sort_values("Time stamp").reset_index(drop=True)
    ps_numeric    = ps_sorted.select_dtypes(include="number").columns.tolist()
    slots, records = get_prescription_slots(merged_sorted), []

    for slot in slots:
        window = merged_sorted[merged_sorted["End Time stamp"] <= slot].copy()
        if window.empty:
            continue
        collected, boxes_needed = [], n
        for idx in reversed(window.index):
            row      = window.loc[idx]
            row_boxes = row["Total Boxes"]
            if row_boxes < 0:
                continue
            if row_boxes <= boxes_needed:
                rd = {c: row[c] for c in SUM_COLS}
                rd.update({"Total Boxes": row_boxes,
                           "Start Time stamp": row["Start Time stamp"],
                           "End Time stamp":   row["End Time stamp"],
                           "_fraction": 1.0})
                collected.append(rd)
                boxes_needed -= row_boxes
            else:
                frac  = boxes_needed / row_boxes
                split = {c: row[c] * frac for c in SUM_COLS}
                span  = row["End Time stamp"] - row["Start Time stamp"]
                split.update({"Total Boxes": boxes_needed,
                              "Start Time stamp": (row["End Time stamp"] - span * frac).floor("min"),
                              "End Time stamp":   row["End Time stamp"],
                              "_fraction": frac})
                collected.append(split)
                boxes_needed = 0
            if boxes_needed <= 0:
                break
        if not collected:
            continue
        cdf          = pd.DataFrame(collected)
        window_start = cdf["Start Time stamp"].min()
        agg          = {c: cdf[c].sum() for c in SUM_COLS}
        ps_win       = ps_sorted[(ps_sorted["Time stamp"] >= window_start) &
                                  (ps_sorted["Time stamp"] <= slot)]
        if ps_win.empty:
            prior  = ps_sorted[ps_sorted["Time stamp"] <= slot]
            ps_win = prior.tail(1) if not prior.empty else pd.DataFrame()
        if ps_win.empty and not ps_sorted.empty:
            ps_win = ps_sorted.head(1)
        ps_avg = ps_win[ps_numeric].mean().to_dict() if not ps_win.empty else {c: None for c in ps_numeric}
        records.append({"Prescription Time": slot,
                        "Window Start": window_start,
                        "Window End":   cdf["End Time stamp"].max(),
                        "Actual Boxes": cdf["Total Boxes"].sum(),
                        **agg, **ps_avg})
    return pd.DataFrame(records)


def run_prescription(ps_path, cons_path, comp_path):
    ps_data  = pd.read_excel(ps_path,   skiprows=5)
    cons_df  = pd.read_excel(cons_path,  skiprows=5)
    comp_df  = pd.read_excel(comp_path,  skiprows=5)

    # Preprocess prepared sand
    ps_data = keep_last_value_only(ps_data[config2["ps_columns"][0]])
    ps_data = ps_data.dropna(subset=["Date", "Time"]).copy()
    ps_data["Time stamp"] = pd.to_datetime(
        ps_data["Date"].astype(str) + " " + ps_data["Time"].astype(str), format="%Y-%m-%d %H:%M"
    )

    cons_df.columns = cons_df.columns.str.strip()
    comp_df.columns = comp_df.columns.str.strip()

    merged = pd.merge(cons_df, comp_df[config2["comp_columns"][0]], on="Component ID", how="left")
    merged["Types of Cores"] = merged["Types of Cores"].astype(str).str.strip().str.title()

    mask_no_core = merged["Types of Cores"].isin(["", "Nan", "None"])
    merged.loc[mask_no_core, config2["core_influx"]] = merged[config2["core_influx"]].fillna(0)
    merged.loc[mask_no_core & (merged[config2["core_influx"]] <= 0), "Types of Cores"] = "No Core"
    merged.loc[mask_no_core & (merged[config2["core_influx"]] >  0), "Types of Cores"] = "Cold Box"
    merged["Core Sand Lost (%)"] = merged["Core Sand Lost (%)"].fillna(20).astype(float)
    merged.loc[merged["Types of Cores"] == "No Core", "Core Sand Lost (%)"] = 20
    merged["Core Influx Sand (MT)"] = merged["Core Sand (MT)"] * ((100 - merged["Core Sand Lost (%)"]) / 100)
    merged["Start Time stamp"] = pd.to_datetime(merged["Date"].astype(str) + " " + merged["Start Time"].astype(str), format="%Y-%m-%d %H:%M")
    merged["End Time stamp"]   = pd.to_datetime(merged["Date"].astype(str) + " " + merged["End Time"].astype(str),   format="%Y-%m-%d %H:%M")
    merged["Total Boxes"]      = merged["No of Boxes"] + merged["Unpoured Moulds (Nos)"]

    prescription_df = build_prescription_records(merged, ps_data)

    ps_global_mean = ps_data.select_dtypes(include="number").mean()
    for col in ps_global_mean.index:
        if col in prescription_df.columns:
            prescription_df[col] = prescription_df[col].fillna(ps_global_mean[col])

    dat = prescription_df.copy()
    dat["sand_metal_ratio"] = dat["Total Prepared Sand (MT)"]   / dat["Total Liquid Metal Poured (MT)"]
    dat["core_sand_ratio"]  = dat["Core Influx Sand (MT)"]       / dat["Total Prepared Sand (MT)"]
    dat["core_metal_ratio"] = dat["Core Influx Sand (MT)"]       / dat["Total Liquid Metal Poured (MT)"]
    dat = dat.rename(columns=config2["ps_data_column_rename"])

    ps_prop_cols = [c for c in config2["props_column"] if c in dat.columns]
    dat[ps_prop_cols] = dat[ps_prop_cols].ffill().bfill()

    required = [c for c in config2["props_column"] + config2["ratio_list"] if c in dat.columns]
    dat.dropna(subset=required, how="any", inplace=True)

    dat[["activeClay","compactibility","gcs","gfnAfs","inertFines","loi",
         "moisture","permeability","shearStrength","splitStrength",
         "tempOfSandAfterMix","volatileMatter"]] = \
        dat[["activeClay","compactibility","gcs","gfnAfs","inertFines","loi",
             "moisture","permeability","shearStrength","splitStrength",
             "tempOfSandAfterMix","volatileMatter"]].round(2)

    dat.reset_index(drop=True, inplace=True)
    dat_copy = dat.copy()

    optimum          = pd.Series(config2["Group_opt"][config2["group_name"]])
    mixture_capacity = config2["batch_size"]
    input_fields     = sorted(config2["additives_list"] + ["return sand_frac"])

    pred, pred_cols = additives_simulation(
        dat_copy, input_fields, model_info[str(config2["real_time_runner_model_no"])],
        optimum, mixture_capacity
    )

    pred_dict = {}
    for i in config2["uncertainity_parameter"]:
        if config2["uncertainity_parameter"][i]:
            tmp = dat.copy()
            for j in config2["uncertainity_parameter"][i]:
                tmp[j] = config2["Group_opt"][config2["group_name"]][j]
            p1, pc1 = additives_simulation(tmp, input_fields, model_info[str(config2["real_time_runner_model_no"])], optimum, mixture_capacity)
            pred_dict[i] = p1

    dat[pred_cols] = pred[pred_cols]
    dat = dat[["Prescription Time"] + config2["props_column"] + config2["ratio_list"] + pred_cols + ["csi"]]
    for i in pred_dict:
        dat[i + "_predicted_kg1"] = pred_dict[i][i + "_predicted_kg"]

    dat.sort_values("Prescription Time", ascending=True, inplace=True)

    # Save prescription file
    presc_path = os.path.join(analysis_dir, "dummy3.xlsx")
    dat.to_excel(presc_path, index=False)

    # Get last row for master update + Final Prediction
    last_row = dat.iloc[-1]
    final_pred = compute_final_prediction(last_row.to_dict())

    # Write C:Y into master Final.xlsx row 41
    wb  = load_workbook(master_path)
    ws  = wb.active

    ws.cell(3, col2idx("Y")).value = n
    ws.cell(4, col2idx("Y")).value = perShift
    ws.cell(1, col2idx("Z")).value = OFFSET

    COL_MAP = {
        "C": "activeClay",         "D": "compactibility",
        "E": "gcs",                "F": "gfnAfs",
        "G": "inertFines",         "H": "loi",
        "I": "moisture",           "J": "permeability",
        "K": "shearStrength",      "L": "splitStrength",
        "M": "tempOfSandAfterMix", "N": "volatileMatter",
        "O": "sand_metal_ratio",   "P": "core_sand_ratio",
        "Q": "core_metal_ratio",   "R": "bentonite_predicted_kg",
        "S": "fresh silica sand_predicted_kg",
        "T": "lca_predicted_kg",   "U": "return sand_predicted_kg",
        "V": "water_predicted_kg", "W": "csi",
        "X": None,                 "Y": "bentonite_predicted_kg1",
    }

    target_row = 6
    for r in range(7, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and not isinstance(v, str):
            target_row = r

    for col_letter, presc_col in COL_MAP.items():
        if presc_col is None:
            ws.cell(target_row, col2idx(col_letter)).value = 0
        else:
            val = last_row.get(presc_col)
            ws.cell(target_row, col2idx(col_letter)).value = val if val is not None and pd.notna(val) else None

    wb.save(master_path)

    def fmt(v):
        try:    return round(float(v), 2)
        except: return v

    results = {
        "activeClay":         fmt(last_row.get("activeClay")),
        "compactibility":     fmt(last_row.get("compactibility")),
        "gcs":                fmt(last_row.get("gcs")),
        "gfnAfs":             fmt(last_row.get("gfnAfs")),
        "inertFines":         fmt(last_row.get("inertFines")),
        "loi":                fmt(last_row.get("loi")),
        "moisture":           fmt(last_row.get("moisture")),
        "permeability":       fmt(last_row.get("permeability")),
        "shearStrength":      fmt(last_row.get("shearStrength")),
        "splitStrength":      fmt(last_row.get("splitStrength")),
        "tempOfSandAfterMix": fmt(last_row.get("tempOfSandAfterMix")),
        "volatileMatter":     fmt(last_row.get("volatileMatter")),
        "csi":                fmt(last_row.get("csi")),
        "smr":                fmt(last_row.get("sand_metal_ratio")),
        "finalPrediction":    final_pred,
    }
    return results


# ── Image generation ──────────────────────────────────────────────────────────
def save_prescription_image(results, output_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    ROWS = [
        ("Active Clay (%)",              results["activeClay"]),
        ("Compactibility (%)",           results["compactibility"]),
        ("GCS (gm/cm²)",                 results["gcs"]),
        ("GFN / AFS (no)",               results["gfnAfs"]),
        ("Inert Fines (%)",              results["inertFines"]),
        ("LOI (%)",                      results["loi"]),
        ("Moisture (%)",                 results["moisture"]),
        ("Permeability (no)",            results["permeability"]),
        ("Shear Strength (gm/cm²)",      results["shearStrength"]),
        ("Split Strength (gm/cm²)",      results["splitStrength"]),
        ("Temp. of Sand After Mix (°C)", results["tempOfSandAfterMix"]),
        ("Volatile Matter (%)",          results["volatileMatter"]),
    ]
    ROW_H, COL_W = 0.44, [2.4, 1.4]
    ROW_CLR, BORDER_CLR, TEXT_CLR = "#FFDAB9", "#C0C0C0", "#1a1a1a"
    total_rows = len(ROWS)
    fig, ax = plt.subplots(figsize=(sum(COL_W) + 0.2, total_rows * ROW_H + 0.3))
    ax.set_xlim(0, sum(COL_W)); ax.set_ylim(0, total_rows * ROW_H); ax.axis("off")

    for i, (label, val) in enumerate(ROWS):
        y  = (total_rows - i - 1) * ROW_H
        x1 = COL_W[0]; x2 = sum(COL_W)
        for xs, xe in [(0, x1), (x1, x2)]:
            ax.add_patch(mpatches.FancyBboxPatch(
                (xs, y), xe - xs, ROW_H,
                boxstyle="square,pad=0", linewidth=0.5,
                edgecolor=BORDER_CLR, facecolor=ROW_CLR))
        ax.text(0.1,  y + ROW_H / 2, label, va="center", ha="left",  fontsize=9, fontweight="bold",   color=TEXT_CLR)
        ax.text(x2 - 0.1, y + ROW_H / 2, str(val) if val is not None else "-", va="center", ha="right", fontsize=9, color=TEXT_CLR)

    plt.tight_layout(pad=0)
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ── Email ─────────────────────────────────────────────────────────────────────
def send_email(image_path, results):
    group_display  = config2["group_name"].replace("_", " ").title()
    new_sand_val   = round(((newsand - 52) * 8.4) / 70, 2)
    new_sand_line  = f"New Sand = (({newsand} - 52) * 8.4) / 70 = {new_sand_val}"

    with open(image_path, "rb") as f:
        img_data = f.read()

    html = f"""
    <html><body>
      <img src="cid:prescription_img" style="display:block;"><br>
      <pre style="font-family:monospace;font-size:13px;">
Last {n} Boxes Total CSI  :  {results['csi']}
SMR                       :  {results['smr']}
{group_display} Bentonite Pred :  {results['finalPrediction']}
{new_sand_line}
      </pre>
    </body></html>
    """
    msg = MIMEMultipart("related")
    SENDER       = os.environ.get("EMAIL_SENDER",   "gokulramesh033@gmail.com")
    APP_PASSWORD = os.environ.get("EMAIL_PASSWORD", "kisesrobrlsjrkds")
    RECIPIENT    = os.environ.get("EMAIL_RECIPIENT","gokul@mpminfosoft.com")

    msg["From"]    = SENDER
    msg["To"]      = RECIPIENT
    msg["Subject"] = "Prescription"
    msg.attach(MIMEText(html, "html"))
    img_mime = MIMEImage(img_data)
    img_mime.add_header("Content-ID", "<prescription_img>")
    img_mime.add_header("Content-Disposition", "inline")
    msg.attach(img_mime)

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(SENDER, APP_PASSWORD)
        server.sendmail(SENDER, RECIPIENT, msg.as_string())


# ── Flask app ─────────────────────────────────────────────────────────────────
app = Flask(__name__)

UPLOAD_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Prescription</title>
  <style>
    body { font-family: Arial, sans-serif; max-width: 500px; margin: 40px auto; padding: 0 20px; background: #f5f5f5; }
    h2   { color: #333; }
    .card { background: white; padding: 24px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
    label { display: block; margin: 14px 0 4px; font-weight: bold; color: #555; }
    input[type=file] { width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 6px; background: #fafafa; }
    button { margin-top: 20px; width: 100%; padding: 14px; background: #F4A460; border: none; border-radius: 8px; font-size: 16px; font-weight: bold; cursor: pointer; color: white; }
    button:hover { background: #e0935a; }
    .result { margin-top: 20px; background: #fff8f0; border: 1px solid #F4A460; border-radius: 8px; padding: 16px; }
    .result table { width: 100%; border-collapse: collapse; }
    .result td { padding: 6px 10px; border-bottom: 1px solid #f0d8c0; }
    .result td:first-child { font-weight: bold; color: #555; }
    .result td:last-child { text-align: right; color: #333; }
    .summary { margin-top: 12px; font-family: monospace; font-size: 13px; background: #fff; padding: 10px; border-radius: 6px; border: 1px solid #eee; }
    .error { color: red; margin-top: 16px; }
    .success { color: green; font-weight: bold; margin-top: 8px; }
  </style>
</head>
<body>
  <div class="card">
    <h2>Prescription Generator</h2>
    <a href="/settings" style="font-size:13px;color:#888;float:right;margin-top:-30px;">⚙ Settings</a>
    <form method="POST" enctype="multipart/form-data">
      <label>Prepared Sand File</label>
      <input type="file" name="preparedsand" accept=".xlsx" required>
      <label>Consumption File</label>
      <input type="file" name="consumption" accept=".xlsx" required>
      <label>Consumption Booking File</label>
      <input type="file" name="consumptionbooking" accept=".xlsx" required>
      <button type="submit">Generate Prescription</button>
    </form>

    {% if error %}
      <p class="error">{{ error }}</p>
    {% endif %}

    {% if results %}
      <div class="result">
        <table>
          <tr><td>Active Clay (%)</td>              <td>{{ results.activeClay }}</td></tr>
          <tr><td>Compactibility (%)</td>           <td>{{ results.compactibility }}</td></tr>
          <tr><td>GCS (gm/cm²)</td>                <td>{{ results.gcs }}</td></tr>
          <tr><td>GFN / AFS (no)</td>              <td>{{ results.gfnAfs }}</td></tr>
          <tr><td>Inert Fines (%)</td>             <td>{{ results.inertFines }}</td></tr>
          <tr><td>LOI (%)</td>                     <td>{{ results.loi }}</td></tr>
          <tr><td>Moisture (%)</td>                <td>{{ results.moisture }}</td></tr>
          <tr><td>Permeability (no)</td>           <td>{{ results.permeability }}</td></tr>
          <tr><td>Shear Strength (gm/cm²)</td>    <td>{{ results.shearStrength }}</td></tr>
          <tr><td>Split Strength (gm/cm²)</td>    <td>{{ results.splitStrength }}</td></tr>
          <tr><td>Temp. of Sand After Mix (°C)</td><td>{{ results.tempOfSandAfterMix }}</td></tr>
          <tr><td>Volatile Matter (%)</td>         <td>{{ results.volatileMatter }}</td></tr>
        </table>
        <table style="width:100%;margin-top:12px;border-top:2px solid #F4A460;">
          <tr><td style="padding:7px 10px;font-weight:bold;color:#555;">Last {{ n }} Boxes Total CSI</td><td style="padding:7px 10px;text-align:right;color:#333;">{{ results.csi }}</td></tr>
          <tr style="background:#fff8f0;"><td style="padding:7px 10px;font-weight:bold;color:#555;">SMR</td><td style="padding:7px 10px;text-align:right;color:#333;">{{ results.smr }}</td></tr>
          <tr><td style="padding:7px 10px;font-weight:bold;color:#555;">{{ group_display }} Bentonite Pred</td><td style="padding:7px 10px;text-align:right;color:#333;">{{ results.finalPrediction }}</td></tr>
          <tr style="background:#fff8f0;"><td colspan="2" style="padding:7px 10px;font-weight:bold;color:#555;">{{ new_sand_line }}</td></tr>
        </table>
        <p class="success">{{ email_status }}</p>
      </div>
    {% endif %}
  </div>
</body>
</html>
"""


@app.route("/", methods=["GET", "POST"])
def index():
    group_display = config2["group_name"].replace("_", " ").title()
    new_sand_val  = round(((newsand - 52) * 8.4) / 70, 2)
    new_sand_line = f"New Sand = (({newsand} - 52) * 8.4) / 70 = {new_sand_val}"

    if request.method == "GET":
        return render_template_string(UPLOAD_HTML, results=None, error=None,
                                      n=n, group_display=group_display, new_sand_line=new_sand_line, email_status="")

    # Save uploaded files
    try:
        FILE_MAP = {
            "preparedsand":       ("Preparedsand",       None),
            "consumption":        ("Consumption",        None),
            "consumptionbooking": ("Consumptionbooking", None),
        }
        saved = {}
        for field, (prefix, _) in FILE_MAP.items():
            f = request.files.get(field)
            if not f or not f.filename:
                return render_template_string(UPLOAD_HTML, results=None,
                                              error=f"Missing file: {field}",
                                              n=n, group_display=group_display,
                                              new_sand_line=new_sand_line, email_status="")
            dest = os.path.join(data_dir, f.filename)
            f.save(dest)
            saved[field] = dest

        # Step 1 — merge consumption + booking to add Start/End Time
        process_consumption_file(saved["consumption"], saved["consumptionbooking"])

        # Step 2 — find Component file (static, stays on server)
        comp_files = glob.glob(os.path.join(data_dir, "Component_*.xlsx"))
        if not comp_files:
            return render_template_string(UPLOAD_HTML, results=None,
                                          error="Component file not found on server.",
                                          n=n, group_display=group_display,
                                          new_sand_line=new_sand_line)

        # Step 3 — run prescription pipeline
        results = run_prescription(
            ps_path   = saved["preparedsand"],
            cons_path = saved["consumption"],
            comp_path = comp_files[0],
        )

        # Step 4 — generate image and send email
        image_path = os.path.join(analysis_dir, "prescription_output.png")
        save_prescription_image(results, image_path)
        email_status = "Email sent to gokul@mpminfosoft.com"
        try:
            send_email(image_path, results)
        except Exception as mail_err:
            email_status = f"Email failed: {mail_err}"

        return render_template_string(UPLOAD_HTML, results=results, error=None,
                                      n=n, group_display=group_display,
                                      new_sand_line=new_sand_line,
                                      email_status=email_status)

    except Exception as e:
        return render_template_string(UPLOAD_HTML, results=None, error=str(e),
                                      n=n, group_display=group_display,
                                      new_sand_line=new_sand_line, email_status="")


SETTINGS_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Settings</title>
  <style>
    body  { font-family: Arial, sans-serif; max-width: 420px; margin: 40px auto; padding: 0 20px; background: #f5f5f5; }
    .card { background: white; padding: 24px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
    h2    { color: #333; margin-top: 0; }
    label { display: block; margin: 14px 0 4px; font-weight: bold; color: #555; }
    input[type=number] { width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 15px; box-sizing: border-box; }
    .hint  { font-size: 12px; color: #999; margin-top: 2px; }
    button { margin-top: 22px; width: 100%; padding: 14px; background: #F4A460; border: none; border-radius: 8px; font-size: 16px; font-weight: bold; cursor: pointer; color: white; }
    button:hover { background: #e0935a; }
    .success { color: green; font-weight: bold; margin-top: 14px; text-align: center; }
    .back { display: block; text-align: center; margin-top: 14px; color: #888; text-decoration: none; font-size: 14px; }
  </style>
</head>
<body>
  <div class="card">
    <h2>Settings</h2>
    <form method="POST">
      <label>n — Total Boxes</label>
      <input type="number" name="n" value="{{ s.n }}" step="1" required>
      <p class="hint">Number of boxes used for prescription window</p>

      <label>perShift — Boxes per Shift</label>
      <input type="number" name="perShift" value="{{ s.perShift }}" step="1" required>
      <p class="hint">Expected boxes per shift</p>

      <label>OFFSET</label>
      <input type="number" name="OFFSET" value="{{ s.OFFSET }}" step="0.1" required>
      <p class="hint">Written to cell Z1 in Final.xlsx</p>

      <label>New Sand Value</label>
      <input type="number" name="newsand" value="{{ s.newsand }}" step="0.1" required>
      <p class="hint">Used in formula: ((newsand - 52) * 8.4) / 70</p>

      <button type="submit">Save Settings</button>
    </form>
    {% if saved %}
    <p class="success">Settings saved successfully.</p>
    {% endif %}
    <a class="back" href="/">← Back to Upload</a>
  </div>
</body>
</html>
"""


@app.route("/settings", methods=["GET", "POST"])
def settings():
    global n, perShift, OFFSET, newsand
    saved = False
    if request.method == "POST":
        new_s = {
            "n":        int(request.form["n"]),
            "perShift": int(request.form["perShift"]),
            "OFFSET":   float(request.form["OFFSET"]),
            "newsand":  float(request.form["newsand"]),
        }
        save_settings(new_s)
        n        = new_s["n"]
        perShift = new_s["perShift"]
        OFFSET   = new_s["OFFSET"]
        newsand  = new_s["newsand"]
        saved    = True
    return render_template_string(SETTINGS_HTML, s=load_settings(), saved=saved)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
