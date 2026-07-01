import os
from preprocess import load_all_configs,read_data,perform_modeling,calculate_correlations,without_bound,dict_conv,save_dict_to_excel,convert_dict,save_dict_to_excel_single_sheet,save_model_info_to_excel
import pandas as pd
import json
from API.helper import additives_simulation
import warnings
from datetime import datetime, time,timedelta
from ps_etl import process_shift_data
from cad_etl import calculate_data_for_all_shifts
warnings.filterwarnings("ignore")
from openpyxl import load_workbook
from openpyxl import Workbook

configs=load_all_configs()    
config1 = configs.get("config.json")
config2 = configs.get(config1["foundry_name"]+".json")

cwd=os.getcwd()
data_dir = os.path.join(cwd,"Data",config1["foundry_name"])
ressults_dir = os.path.join(cwd,"Model Results",config1["foundry_name"])
analysis_dir = os.path.join(cwd,"Analysis",config1["foundry_name"])

for file_name in os.listdir(data_dir):
    if file_name.startswith("Preparedsand") and file_name.endswith(".xlsx"):
            file_path = os.path.join(data_dir, file_name)
            ps_data = pd.read_excel(file_path, skiprows=5)
    elif file_name.startswith("Consumption") and file_name.endswith(".xlsx"):
            file_path = os.path.join(data_dir, file_name)
            cons_df = pd.read_excel(file_path, skiprows=5)
    elif file_name.startswith("Component") and file_name.endswith(".xlsx"):
            file_path = os.path.join(data_dir, file_name)
            comp_df = pd.read_excel(file_path, skiprows=5) 


test_data = True
with open(os.path.join(ressults_dir,'output2_props.json'), 'r') as json_file:
    model_info = json.load(json_file)

if test_data:
    ps_data = ps_data[config2['ps_columns'][0]]

    ps_data = ps_data .groupby(["Date","Shift"],as_index=False).mean()

    cons_df.columns = cons_df.columns.str.strip()
    comp_df.columns = comp_df.columns.str.strip()


    merged = pd.merge(
        cons_df,
        comp_df[config2['comp_columns'][0]],
        on='Component ID',
        how='left'
    )

    merged['Types of Cores'] = merged['Types of Cores'].astype(str).str.strip().str.title()

    mask_no_core_info = merged['Types of Cores'].isin(['', 'Nan', 'None'])
    merged.loc[mask_no_core_info, config2["core_influx"]] = merged[config2["core_influx"]].fillna(0)

    merged.loc[
        mask_no_core_info & (merged[config2["core_influx"]] <= 0),
        'Types of Cores'
    ] = 'No Core'

    merged.loc[
        mask_no_core_info & (merged[config2["core_influx"]] > 0),
        'Types of Cores'
    ] = 'Cold Box'

    merged['Core Sand Lost (%)'] = merged['Core Sand Lost (%)'].fillna(20)
    merged.loc[merged['Types of Cores'] == 'No Core', 'Core Sand Lost (%)'] = 20 

    merged['Core Sand Lost (%)'] = merged['Core Sand Lost (%)'].astype(float)
    merged['Core Influx Sand (MT)'] = merged['Core Sand (MT)'] * ((100 - merged['Core Sand Lost (%)']) / 100)

    dat = merged.groupby(['Date', 'Shift'], as_index=False)[
        ['Core Sand (MT)', 'Total Prepared Sand (MT)', 'Total Liquid Metal Poured (MT)', 'Core Influx Sand (MT)']
    ].sum()

    dat['sand_metal_ratio'] = dat['Total Prepared Sand (MT)'] / dat['Total Liquid Metal Poured (MT)']
    dat['core_sand_ratio'] = dat['Core Influx Sand (MT)'] / dat['Total Prepared Sand (MT)']
    dat['core_metal_ratio'] = dat['Core Influx Sand (MT)'] / dat['Total Liquid Metal Poured (MT)']
    # dat[['Date', 'Shift', 'Core Influx Sand (MT)', 'SMR','CSR','CMR']].to_excel("Cons_mean.xlsx", index=False)
    test_data = pd.merge(left = dat,right = ps_data,how = "left",on = ["Date","Shift"])
    exclude_cols = ['Date', 'Shift']

# Columns to shift
    cols_to_shift = [col for col in test_data.columns if col not in exclude_cols]
    test_data[cols_to_shift] = test_data[cols_to_shift].shift(1)
    test_data = test_data.ffill().bfill()
    test_data=test_data.rename(columns=config2["ps_data_column_rename"])
else:
    data = pd.read_excel(os.path.join(data_dir,"Test Data","Preparedsand_01-Jan-2026_TO_06-Mar-2026_SAVELLI.xlsx"),skiprows=5)
    data_1=data.copy()
    ps_data=data.groupby(['Date','Shift'].mean())
    # ps_data.to_excel("temp_ps.xlsx")
    ps_data.rename(columns=config2["ps_data_column_rename"],inplace=True)
    ps_data.dropna(axis=1,how="all",inplace=True)
    cs_data = pd.read_excel(os.path.join(data_dir,"Test Data","Consumption_01-Jan-2026_TO_06-Mar-2026_SAVELLI.xlsx"),skiprows=5)
    data_2=cs_data.copy()
    cons_data=data_2.groupby(['Date','Shift'].sum())
    cs_data['Date and Shift'] = cs_data["Date"].astype(str)+" "+cs_data["Shift"].astype(str)
    last_date_shift=cs_data["Date and Shift"].max()
    print("Total no of boxes in",last_date_shift,"is",cs_data.loc[cs_data["Date and Shift"]==last_date_shift,"No of Boxes"].sum())
   
    cons_data=calculate_data_for_all_shifts(data_2,config2)
    # cons_data.to_excel("temp_cons.xlsx")
    # cons_data.to_excel("test_Data.xlsx",index = False)
    cons_data.rename(columns=config2["cons_data_column_rename"],inplace=True)
    cons_data.dropna(axis=1,how="all",inplace=True)
    ps_data_agg = ps_data.groupby(["date","shift"],as_index=False).mean()
    # ps_data_agg.loc[ps_data_agg["shift"]>="Shift_24","date"]=ps_data_agg.loc[ps_data_agg["shift"]>="Shift_24","date"]-timedelta(days =1)
    cons_data_agg = cons_data.groupby(["date","shift"],as_index=False).sum()
    cons_data_agg["csi"] = cons_data_agg["csi"]*0.65
    cons_data_agg["core_metal_ratio"]=cons_data_agg["csi"]/cons_data_agg["tlm"]
    cons_data_agg["sand_metal_ratio"]=cons_data_agg["tps"]/cons_data_agg["tlm"]
    cons_data_agg["core_sand_ratio"]=cons_data_agg["csi"]/cons_data_agg["tps"]
    test_data = pd.merge(left = ps_data_agg,right=cons_data_agg,on=["date","shift"],how = "inner")
    test_data = test_data[["date","shift"]+config2["props_column"]+config2["ratio_list"]+["csi","fss"]]

    # test_data[config2["props_column"]] = test_data[config2["props_column"]].ffill().bfill()
    test_data.dropna(how="any",inplace=True,axis=0)

test_data[["activeClay","compactibility","gcs","gfnAfs","inertFines","loi","moisture","permeability","shearStrength","splitStrength","tempOfSandAfterMix","volatileMatter"]] = test_data[["activeClay","compactibility","gcs","gfnAfs","inertFines","loi","moisture","permeability","shearStrength","splitStrength","tempOfSandAfterMix","volatileMatter"]].round(2)

test_data.reset_index(inplace=True,drop=True)
test_data = test_data.sort_index(axis=1)
test_data_copy = test_data.drop(["date","shift"],axis=1)

# test_data_copy.to_excel("test_data.xlsx")

# for cols in test_data.columns:
#     min_val = model_info[str(config2["real_time_runner_model_no"])]["model_info"]["input"][cols]['min']
#     max_val = model_info[str(config2["real_time_runner_model_no"])]["model_info"]["input"][cols]['max']
#     test_data_copy[cols] = (test_data_copy[cols] - min_val) / (max_val - min_val)
#     mean_val = model_info[str(config2["real_time_runner_model_no"])]["model_info"]["input"][cols]['mean']
#     test_data_copy[cols] = test_data_copy[cols] - mean_val

optimum = pd.Series(config2['Group_opt'][config2["group_name"]])
Mixture_capacity = config2["batch_size"]

inputFieldsList2 = config2["additives_list"]+["return sand_frac"]
inputFieldsList2.sort()

# for i in inputFieldsList2:
#     test_data_copy[i]=0

predicted_additive, predicted_additive_column = additives_simulation(test_data_copy, inputFieldsList2, model_info[str(config2["real_time_runner_model_no"])], optimum, Mixture_capacity)

predicted_additive_dict = {}

for i in config2["uncertainity_parameter"]:
    if len(config2["uncertainity_parameter"][i])>0:
        test_data_copy = test_data.drop(["date","shift"],axis=1)
        for j in config2["uncertainity_parameter"][i]:
            test_data_copy[j] = config2["Group_opt"][config2["group_name"]][j]
        predicted_additive1, predicted_additive_column1 = additives_simulation(test_data_copy, inputFieldsList2, model_info[str(config2["real_time_runner_model_no"])], optimum, Mixture_capacity)
        predicted_additive_dict[i] = predicted_additive1


test_data[predicted_additive_column] = predicted_additive[predicted_additive_column]

test_data = test_data[["date","shift"]+config2["props_column"]+config2["ratio_list"]+predicted_additive_column+["csi"]]
# test_data.sort_values(by = ["date","shift"],ascending = True,inplace = True)
# test_data[["date","shift"]] = test_data[["date","shift"]].shift(-1)

for i in predicted_additive_dict:
    test_data[i+"_predicted_kg1"] = predicted_additive_dict[i][i+"_predicted_kg"]


test_data.sort_values(by = ["date","shift"],ascending = True,inplace = True)
#test_data[["date","shift"]] = test_data[["date","shift"]].shift(-1)
test_data.to_excel(os.path.join(analysis_dir,"dummy_GPI_2.xlsx"),index=False)



# output_path = "test_data_gpi.xlsx"
# test_data.to_excel(output_path, startrow=6, index=False)

# wb = load_workbook(output_path)
# ws = wb.active

# row_start = 1

# # Extract configs
# group_opt = config2["Group_opt"]["group_2"]
# adjustment = config2["Adjustment"]["group_2"]

# # Step 1: All parameters (fixed order)
# all_params = list(group_opt.keys())

# # Step 2: Target Values
# ws.cell(row=1, column=1, value="Target Values")
# for col, param in enumerate(all_params, start=2):
#     ws.cell(row=1, column=col, value=group_opt[param])

# # Step 3: Initialize coeff & weight dict
# coeff_dict = {p: 9999 for p in all_params}
# weight_dict = {p: 0 for p in all_params}

# # Step 4: Fill from Adjustment
# for material, values in adjustment.items():
#     params = values["param"]
#     coeffs = values["ref_coeff"]
#     weights = values["weight"]
    
#     for p, c, w in zip(params, coeffs, weights):
#         coeff_dict[p] = c
#         weight_dict[p] = w

# # Step 5: Write Coefficients
# ws.cell(row=2, column=1, value="Coefficient")
# for col, param in enumerate(all_params, start=2):
#     ws.cell(row=2, column=col, value=coeff_dict[param])

# # Step 6: Write Weights
# ws.cell(row=3, column=1, value="Weight")
# for col, param in enumerate(all_params, start=2):
#     ws.cell(row=3, column=col, value=weight_dict[param])

# # Save file
# wb.save("group_2_final_table.xlsx")