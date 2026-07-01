import pandas as pd
from openpyxl import load_workbook
import os
import glob

DATA_DIR = r"C:\Users\gokul\Downloads\Final (5) 14\Final (5)\Data\GPI"

SKIPROWS = 5

# Only the last existing value is kept; all earlier values in these columns are cleared
LAST_VALUE_ONLY_COLS = [
    "Active Clay (%)",
    "GFN/AFS (no)",
    "Inert Fines (%)",
    "LOI (%)",
    "Volatile Matter (%)",
]


def keep_last_value_only(df):
    """
    For each column in LAST_VALUE_ONLY_COLS, find the last non-null row
    and set every other row in that column to NaN.
    All other columns are left untouched.
    """
    df = df.copy()
    for col in LAST_VALUE_ONLY_COLS:
        if col not in df.columns:
            continue
        non_null_idx = df.index[df[col].notna()].tolist()
        if len(non_null_idx) <= 1:
            continue
        # Clear every non-null row except the last one
        rows_to_clear = non_null_idx[:-1]
        df.loc[rows_to_clear, col] = None
    return df


def preprocess_prepared_sand(filepath):
    df = pd.read_excel(filepath, skiprows=SKIPROWS)
    df = keep_last_value_only(df)
    write_back_excel(filepath, df, original_header_rows=SKIPROWS)
    print(f"Processed: {filepath}")


def write_back_excel(filepath, df, original_header_rows=5):
    """
    Overwrite the data portion of an Excel file (below the original header rows)
    while preserving the top header rows exactly as-is.
    """
    wb = load_workbook(filepath)
    ws = wb.active

    if ws.max_row > original_header_rows:
        ws.delete_rows(original_header_rows + 1, ws.max_row - original_header_rows)

    header_row = original_header_rows + 1
    col_names = list(df.columns)
    for col_idx, col_name in enumerate(col_names, 1):
        ws.cell(row=header_row, column=col_idx, value=col_name)

    date_col_idx = col_names.index("Date") + 1 if "Date" in col_names else None

    for row_offset, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell_value = value
            if pd.isnull(value) if not isinstance(value, str) else False:
                cell_value = None
            cell = ws.cell(row=header_row + row_offset - 1, column=col_idx, value=cell_value)

            if col_idx == date_col_idx and cell_value is not None:
                cell.number_format = "DD-MM-YYYY"

    wb.save(filepath)


if __name__ == "__main__":
    pattern = os.path.join(DATA_DIR, "Preparedsand_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        print(f"No Preparedsand files found in {DATA_DIR}")
    for f in files:
        preprocess_prepared_sand(f)
