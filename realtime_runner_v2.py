import os
from preprocess import load_all_configs,read_data,perform_modeling,calculate_correlations,without_bound,dict_conv,save_dict_to_excel,convert_dict,save_dict_to_excel_single_sheet,save_model_info_to_excel
import pandas as pd
import json
from API.helper import additives_simulation
import warnings
from datetime import datetime, time,timedelta
from ps_etl import process_shift_data
from cad_etl import calculate_data_for_all_shifts
warnings.filterwarnings("ignore")
from openpyxl import load_workbook
from openpyxl import Workbook
from typing import List

configs=load_all_configs()
config1 = configs.get("config.json")
config2 = configs.get(config1["foundry_name"]+".json")
PRESCRIPTION_TIMES = [time(0, 0), time(8, 0), time(16, 0)]
n        = 1500
perShift = 450
OFFSET   = 5


newsand = 42


cwd=os.getcwd()
data_dir = os.path.join(cwd,"Data",config1["foundry_name"])
ressults_dir = os.path.join(cwd,"Model Results",config1["foundry_name"])
analysis_dir = os.path.join(cwd,"Analysis",config1["foundry_name"])

for file_name in os.listdir(data_dir):
    if file_name.startswith("Preparedsand") and file_name.endswith(".xlsx"):
            file_path = os.path.join(data_dir, file_name)
            ps_data = pd.read_excel(file_path, skiprows=5)
    elif file_name.startswith("Consumption") and file_name.endswith(".xlsx"):
            file_path = os.path.join(data_dir, file_name)
            cons_df = pd.read_excel(file_path, skiprows=5)
    elif file_name.startswith("Component") and file_name.endswith(".xlsx"):
            file_path = os.path.join(data_dir, file_name)
            comp_df = pd.read_excel(file_path, skiprows=5)


test_data = True
with open(os.path.join(ressults_dir,'output2_props.json'), 'r') as json_file:
    model_info = json.load(json_file)

if test_data:
    ps_data = ps_data[config2['ps_columns'][0]]

    ps_data = ps_data.dropna(subset=["Date", "Time"]).copy()
    ps_data["Time stamp"] = pd.to_datetime(ps_data["Date"].astype(str)+" "+ps_data["Time"].astype(str), format="%Y-%m-%d %H:%M")

    cons_df.columns = cons_df.columns.str.strip()
    comp_df.columns = comp_df.columns.str.strip()


    merged = pd.merge(
        cons_df,
        comp_df[config2['comp_columns'][0]],
        on='Component ID',
        how='left'
    )

    merged['Types of Cores'] = merged['Types of Cores'].astype(str).str.strip().str.title()

    mask_no_core_info = merged['Types of Cores'].isin(['', 'Nan', 'None'])
    merged.loc[mask_no_core_info, config2["core_influx"]] = merged[config2["core_influx"]].fillna(0)

    merged.loc[
        mask_no_core_info & (merged[config2["core_influx"]] <= 0),
        'Types of Cores'
    ] = 'No Core'

    merged.loc[
        mask_no_core_info & (merged[config2["core_influx"]] > 0),
        'Types of Cores'
    ] = 'Cold Box'

    merged['Core Sand Lost (%)'] = merged['Core Sand Lost (%)'].fillna(20)
    merged.loc[merged['Types of Cores'] == 'No Core', 'Core Sand Lost (%)'] = 20

    merged['Core Sand Lost (%)'] = merged['Core Sand Lost (%)'].astype(float)
    merged['Core Influx Sand (MT)'] = merged['Core Sand (MT)'] * ((100 - merged['Core Sand Lost (%)']) / 100)

    merged["Start Time stamp"] = pd.to_datetime(merged["Date"].astype(str)+" "+merged["Start Time"].astype(str), format="%Y-%m-%d %H:%M")
    merged["End Time stamp"] = pd.to_datetime(merged["Date"].astype(str)+" "+merged["End Time"].astype(str), format="%Y-%m-%d %H:%M")
    merged["Total Boxes"] = merged["No of Boxes"]+merged["Unpoured Moulds (Nos)"]

    # ── Prescription Scheduling ──────────────────────────────────────────────────



    def get_prescription_slots(df: pd.DataFrame) -> List[datetime]:
        """Return all prescription trigger datetimes covering the data range."""
        if df.empty:
            return []
        min_date = df["Start Time stamp"].min().date()
        max_date = df["End Time stamp"].max().date()

        slots = []
        current = min_date
        while current <= max_date:
            for t in PRESCRIPTION_TIMES:
                slots.append(datetime.combine(current, t))
            current += timedelta(days=1)
        return sorted(slots)


    def get_next_slot(dt: datetime) -> datetime:
        """Return the next prescription trigger after dt."""
        for t in PRESCRIPTION_TIMES:
            candidate = datetime.combine(dt.date(), t)
            if candidate > dt:
                return candidate
        # Roll over to midnight of next day
        return datetime.combine(dt.date() + timedelta(days=1), PRESCRIPTION_TIMES[0])


    SUM_COLS = [
        'Core Sand (MT)',
        'Total Prepared Sand (MT)',
        'Total Liquid Metal Poured (MT)',
        'Core Influx Sand (MT)',
    ]


    def build_prescription_records(merged: pd.DataFrame, ps_data: pd.DataFrame, n: int) -> pd.DataFrame:

        merged_sorted = merged.sort_values("End Time stamp").reset_index(drop=True)
        ps_sorted = ps_data.sort_values("Time stamp").reset_index(drop=True)

        ps_numeric_cols = ps_sorted.select_dtypes(include="number").columns.tolist()

        slots = get_prescription_slots(merged_sorted)
        records = []

        for slot in slots:
            window = merged_sorted[merged_sorted["End Time stamp"] <= slot].copy()

            if window.empty:
                continue

            collected_rows = []
            boxes_needed = n

            for idx in reversed(window.index):
                row = window.loc[idx]
                row_boxes = row["Total Boxes"]

                if row_boxes < 0:
                    continue

                if row_boxes <= boxes_needed:
                    row_dict = {col: row[col] for col in SUM_COLS}
                    row_dict["Total Boxes"] = row_boxes
                    row_dict["Start Time stamp"] = row["Start Time stamp"]
                    row_dict["End Time stamp"]   = row["End Time stamp"]
                    row_dict["_fraction"]        = 1.0
                    collected_rows.append(row_dict)
                    boxes_needed -= row_boxes
                else:
                    fraction = boxes_needed / row_boxes
                    split = {col: row[col] * fraction for col in SUM_COLS}
                    split["Total Boxes"] = boxes_needed
                    span = row["End Time stamp"] - row["Start Time stamp"]

                    # ✅ Floor to minute to avoid nanosecond precision artifacts
                    interpolated_start = row["End Time stamp"] - span * fraction
                    split["Start Time stamp"] = interpolated_start.floor("min")
                    split["End Time stamp"]   = row["End Time stamp"]
                    split["_fraction"]        = fraction
                    collected_rows.append(split)
                    boxes_needed = 0

                if boxes_needed <= 0:
                    break

            if not collected_rows:
                continue

            collected_df    = pd.DataFrame(collected_rows)
            actual_boxes    = collected_df["Total Boxes"].sum()
            window_start    = collected_df["Start Time stamp"].min()
            window_end      = collected_df["End Time stamp"].max()

            agg = {col: collected_df[col].sum() for col in SUM_COLS}

            # ✅ Wider PS lookup: use window_start → slot
            ps_window = ps_sorted[
                (ps_sorted["Time stamp"] >= window_start) &
                (ps_sorted["Time stamp"] <= slot)
            ]

            # ✅ Fallback 1: nearest record before slot
            if ps_window.empty:
                prior = ps_sorted[ps_sorted["Time stamp"] <= slot]
                ps_window = prior.tail(1) if not prior.empty else pd.DataFrame()

            # ✅ Fallback 2: nearest record after window_start (data delay tolerance)
            if ps_window.empty:
                after = ps_sorted[ps_sorted["Time stamp"] >= window_start]
                ps_window = after.head(1) if not after.empty else pd.DataFrame()

            # ✅ Fallback 3: use the single available PS record for all slots
            if ps_window.empty and not ps_sorted.empty:
                ps_window = ps_sorted.head(1)

            ps_avg = (
                ps_window[ps_numeric_cols].mean().to_dict()
                if not ps_window.empty
                else {col: None for col in ps_numeric_cols}
            )

            record = {
                "Prescription Time": slot,
                "Window Start":      window_start,
                "Window End":        window_end,
                "Actual Boxes":      actual_boxes,
                **agg,
                **{"" + k: v for k, v in ps_avg.items()},
            }
            records.append(record)

        prescription_df = pd.DataFrame(records)
        return prescription_df


    # ── Run ───────────────────────────────────────────────────────────────────────

    prescription_df = build_prescription_records(merged, ps_data, n)

    # Fill any NaN PS columns using global mean across ALL PS data rows.
    # This ensures that even a single available reading (e.g. one Active Clay value)
    # gets propagated to every prescription record.
    ps_numeric_cols_all = ps_data.select_dtypes(include="number").columns.tolist()
    ps_global_mean = ps_data[ps_numeric_cols_all].mean()
    for col in ps_numeric_cols_all:
        if col in prescription_df.columns:
            prescription_df[col] = prescription_df[col].fillna(ps_global_mean[col])

    print(f"Generated {len(prescription_df)} prescription records")
    print(prescription_df.head())

    dat = prescription_df.copy()

    dat['sand_metal_ratio'] = dat['Total Prepared Sand (MT)'] / dat['Total Liquid Metal Poured (MT)']
    dat['core_sand_ratio'] = dat['Core Influx Sand (MT)'] / dat['Total Prepared Sand (MT)']
    dat['core_metal_ratio'] = dat['Core Influx Sand (MT)'] / dat['Total Liquid Metal Poured (MT)']
    # dat[['Date', 'Shift', 'Core Influx Sand (MT)', 'SMR','CSR','CMR']].to_excel("Cons_mean.xlsx", index=False)
    exclude_cols = ['Date', 'Shift']

# Columns to shift


    dat=dat.rename(columns=config2["ps_data_column_rename"])
    test_data = dat.copy()

    # Forward-fill then back-fill PS property columns so that a single
    # available reading is propagated to all prescription windows.
    ps_prop_cols = [c for c in config2["props_column"] if c in test_data.columns]
    test_data[ps_prop_cols] = test_data[ps_prop_cols].ffill().bfill()

    # Only drop rows where the required model input columns are still missing.
    required_cols = [c for c in config2["props_column"] + config2["ratio_list"] if c in test_data.columns]
    test_data.dropna(subset=required_cols, how="any", inplace=True)

    print("Shape of test Data", test_data.shape)
    print(test_data.columns.tolist())

else:
    data = pd.read_excel(os.path.join(data_dir,"Test Data","Preparedsand_15-Dec-2025_TO_15-Dec-2025_Spomatic.xlsx"),skiprows=5)
    data_1=data.copy()
    ps_data=process_shift_data(data_1,config2)
    # ps_data.to_excel("temp_ps.xlsx")
    ps_data.rename(columns=config2["ps_data_column_rename"],inplace=True)
    ps_data.dropna(axis=1,how="all",inplace=True)
    cs_data = pd.read_excel(os.path.join(data_dir,"Test Data","CS Data(1) 1.xlsx"),skiprows=5)
    data_2=cs_data.copy()
    cs_data['Date and Shift'] = cs_data["Date"].astype(str)+" "+cs_data["Shift"].astype(str)
    last_date_shift=cs_data["Date and Shift"].max()
    print("Total no of boxes in",last_date_shift,"is",cs_data.loc[cs_data["Date and Shift"]==last_date_shift,"No of Boxes"].sum())
    cons_data=calculate_data_for_all_shifts(data_2,config2)
    # cons_data.to_excel("temp_cons.xlsx")
    # cons_data.to_excel("test_Data.xlsx",index = False)
    cons_data.rename(columns=config2["cons_data_column_rename"],inplace=True)
    cons_data.dropna(axis=1,how="all",inplace=True)
    ps_data_agg = ps_data.groupby(["date","shift"],as_index=False).mean()
    # ps_data_agg.loc[ps_data_agg["shift"]>="Shift_24","date"]=ps_data_agg.loc[ps_data_agg["shift"]>="Shift_24","date"]-timedelta(days =1)
    cons_data_agg = cons_data.groupby(["date","shift"],as_index=False).sum()
    cons_data_agg["csi"] = cons_data_agg["csi"]*0.8
    cons_data_agg["core_metal_ratio"]=cons_data_agg["csi"]/cons_data_agg["tlm"]
    cons_data_agg["sand_metal_ratio"]=cons_data_agg["tps"]/cons_data_agg["tlm"]
    cons_data_agg["core_sand_ratio"]=cons_data_agg["csi"]/cons_data_agg["tps"]
    test_data = pd.merge(left = ps_data_agg,right=cons_data_agg,on=["date","shift"],how = "inner")
    test_data = test_data[["date","shift"]+config2["props_column"]+config2["ratio_list"]+["csi","fss"]]

    # test_data[config2["props_column"]] = test_data[config2["props_column"]].ffill().bfill()
    test_data.dropna(how="any",inplace=True,axis=0)

test_data[["activeClay","compactibility","gcs","gfnAfs","inertFines","loi","moisture","permeability","shearStrength","splitStrength","tempOfSandAfterMix","volatileMatter"]] = test_data[["activeClay","compactibility","gcs","gfnAfs","inertFines","loi","moisture","permeability","shearStrength","splitStrength","tempOfSandAfterMix","volatileMatter"]].round(2)

test_data.reset_index(inplace=True,drop=True)
test_data = test_data.sort_index(axis=1)
test_data_copy = test_data.copy()

# test_data_copy.to_excel("test_data.xlsx")

# for cols in test_data.columns:
#     min_val = model_info[str(config2["real_time_runner_model_no"])]["model_info"]["input"][cols]['min']
#     max_val = model_info[str(config2["real_time_runner_model_no"])]["model_info"]["input"][cols]['max']
#     test_data_copy[cols] = (test_data_copy[cols] - min_val) / (max_val - min_val)
#     mean_val = model_info[str(config2["real_time_runner_model_no"])]["model_info"]["input"][cols]['mean']
#     test_data_copy[cols] = test_data_copy[cols] - mean_val

optimum = pd.Series(config2['Group_opt'][config2["group_name"]])
Mixture_capacity = config2["batch_size"]

inputFieldsList2 = config2["additives_list"]+["return sand_frac"]
inputFieldsList2.sort()

# for i in inputFieldsList2:
#     test_data_copy[i]=0

predicted_additive, predicted_additive_column = additives_simulation(test_data_copy, inputFieldsList2, model_info[str(config2["real_time_runner_model_no"])], optimum, Mixture_capacity)

predicted_additive_dict = {}

for i in config2["uncertainity_parameter"]:
    if len(config2["uncertainity_parameter"][i])>0:
        test_data_copy = test_data.copy()
        for j in config2["uncertainity_parameter"][i]:
            test_data_copy[j] = config2["Group_opt"][config2["group_name"]][j]
        predicted_additive1, predicted_additive_column1 = additives_simulation(test_data_copy, inputFieldsList2, model_info[str(config2["real_time_runner_model_no"])], optimum, Mixture_capacity)
        predicted_additive_dict[i] = predicted_additive1


test_data[predicted_additive_column] = predicted_additive[predicted_additive_column]

test_data = test_data[["Prescription Time"]+config2["props_column"]+config2["ratio_list"]+predicted_additive_column+["csi"]]
# test_data.sort_values(by = ["date","shift"],ascending = True,inplace = True)
# test_data[["date","shift"]] = test_data[["date","shift"]].shift(-1)

for i in predicted_additive_dict:
    test_data[i+"_predicted_kg1"] = predicted_additive_dict[i][i+"_predicted_kg"]


test_data.sort_values(by = ["Prescription Time"],ascending = True,inplace = True)
#test_data[["date","shift"]] = test_data[["date","shift"]].shift(-1)
# test_data = test_data[pd.to_datetime(test_data["Prescription Time"]).dt.hour == 16]

prescription_path = os.path.join(analysis_dir, "dummy3.xlsx")
test_data.to_excel(prescription_path, index=False)

# ── Master Workbook Update & Result Extraction ────────────────────────────────

def update_master_and_extract(prescription_path, master_path, header_row=6):
    from openpyxl.utils import column_index_from_string as col2idx

    # Read last row from prescription file
    presc_df = pd.read_excel(prescription_path)
    last_row = presc_df.iloc[-1]

    # Explicit mapping: Excel column letter → prescription column name
    COL_MAP = {
        "C": "activeClay",
        "D": "compactibility",
        "E": "gcs",
        "F": "gfnAfs",
        "G": "inertFines",
        "H": "loi",
        "I": "moisture",
        "J": "permeability",
        "K": "shearStrength",
        "L": "splitStrength",
        "M": "tempOfSandAfterMix",
        "N": "volatileMatter",
        "O": "sand_metal_ratio",
        "P": "core_sand_ratio",
        "Q": "core_metal_ratio",
        "R": "bentonite_predicted_kg",
        "S": "fresh silica sand_predicted_kg",
        "T": "lca_predicted_kg",
        "U": "return sand_predicted_kg",
        "V": "water_predicted_kg",
        "W": "csi",
        "X": None,  # fss is always 0
        "Y": "bentonite_predicted_kg1",
    }

    wb = load_workbook(master_path)
    ws = wb.active

    # Find the last data row (col A holds a non-string, non-None value) and overwrite it
    target_row = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and not isinstance(v, str):
            target_row = r

    # Write code-defined values into Final.xlsx
    ws.cell(3, col2idx("Y")).value = n
    ws.cell(4, col2idx("Y")).value = perShift
    ws.cell(1, col2idx("Z")).value = OFFSET

    # Write values into C:Y using the column letter map
    for col_letter, presc_col in COL_MAP.items():
        if presc_col is None:
            ws.cell(target_row, col2idx(col_letter)).value = 0
        else:
            val = last_row.get(presc_col)
            ws.cell(target_row, col2idx(col_letter)).value = val if val is not None and pd.notna(val) else None

    wb.save(master_path)

    # Open with Excel COM so formulas recalculate automatically
    try:
        import win32com.client
        abs_master = os.path.abspath(master_path)
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb_com = xl.Workbooks.Open(abs_master)
        xl.Calculate()
        wb_com.Save()
        wb_com.Close(SaveChanges=True)
        xl.Quit()
    except Exception as e:
        print(f"Warning: Excel recalculation failed ({e}). Formula values may be stale.")

    # Read back the computed values using column letters directly
    wb_out = load_workbook(master_path, data_only=True)
    ws_out = wb_out.active

    EXTRACT = {
        "W":  "csi",
        "O":  "smr",
        "Z":  "finalPrediction",
        "C":  "activeClay",
        "D":  "compactibility",
        "E":  "gcs",
        "F":  "gfnAfs",
        "G":  "inertFines",
        "H":  "loi",
        "I":  "moisture",
        "J":  "permeability",
        "K":  "shearStrength",
        "L":  "splitStrength",
        "M":  "tempOfSandAfterMix",
        "N":  "volatileMatter",
    }

    results = {}
    for col_letter, key in EXTRACT.items():
        results[key] = ws_out.cell(target_row, col2idx(col_letter)).value

    return results, target_row


master_path = os.path.join(analysis_dir, "Final.xlsx")
prescription_results, appended_row = update_master_and_extract(prescription_path, master_path)

def _fmt(val):
    try:
        return round(float(val), 2)
    except (TypeError, ValueError):
        return val


def save_prescription_image(results, output_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    SAND_ROWS = [
        ("Active Clay (%)",              results.get("activeClay")),
        ("Compactibility (%)",           results.get("compactibility")),
        ("GCS (gm/cm²)",                 results.get("gcs")),
        ("GFN / AFS (no)",               results.get("gfnAfs")),
        ("Inert Fines (%)",              results.get("inertFines")),
        ("LOI (%)",                      results.get("loi")),
        ("Moisture (%)",                 results.get("moisture")),
        ("Permeability (no)",            results.get("permeability")),
        ("Shear Strength (gm/cm²)",      results.get("shearStrength")),
        ("Split Strength (gm/cm²)",      results.get("splitStrength")),
        ("Temp. of Sand After Mix (°C)", results.get("tempOfSandAfterMix")),
        ("Volatile Matter (%)",          results.get("volatileMatter")),
    ]

    ROW_H      = 0.44
    COL_W      = [2.4, 1.4]
    ROW_CLR    = "#FFDAB9"
    BORDER_CLR = "#C0C0C0"
    TEXT_CLR   = "#1a1a1a"

    total_rows = len(SAND_ROWS)
    fig_h = total_rows * ROW_H + 0.3
    fig_w = sum(COL_W) + 0.2

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, sum(COL_W))
    ax.set_ylim(0, total_rows * ROW_H)
    ax.axis("off")

    def draw_row(row_idx, label, value, bg, bold_label=False):
        y = (total_rows - row_idx - 1) * ROW_H
        x0, x1 = 0, COL_W[0]
        x2 = x0 + COL_W[0] + COL_W[1]

        for x_start, x_end in [(x0, x1), (x1, x2)]:
            rect = mpatches.FancyBboxPatch(
                (x_start, y), x_end - x_start, ROW_H,
                boxstyle="square,pad=0", linewidth=0.5,
                edgecolor=BORDER_CLR, facecolor=bg
            )
            ax.add_patch(rect)

        fw = "bold" if bold_label else "bold"
        ax.text(x0 + 0.1, y + ROW_H / 2, label,
                va="center", ha="left", fontsize=9,
                fontweight=fw, color=TEXT_CLR)
        ax.text(x2 - 0.1, y + ROW_H / 2, str(_fmt(value)) if value is not None else "-",
                va="center", ha="right", fontsize=9,
                fontweight="normal", color=TEXT_CLR)

    for i, (label, val) in enumerate(SAND_ROWS):
        draw_row(i, label, val, bg=ROW_CLR)

    plt.tight_layout(pad=0)
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Prescription image saved: {output_path}")


r = prescription_results
image_path = os.path.join(analysis_dir, "prescription_output.png")
save_prescription_image(r, image_path)


def send_prescription_email(image_path, results, boxes, group_name, newsand):
    import smtplib

    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage

    SENDER       = "gokulramesh033@gmail.com"
    APP_PASSWORD = "kisesrobrlsjrkds"
    RECIPIENT    = "gokul@mpminfosoft.com"
    SUBJECT      = "Prescription"

    group_display  = group_name.replace("_", " ").title()
    new_sand_val   = _fmt(((newsand - 52) * 8.4) / 70)
    new_sand_line  = f"New Sand = (({newsand} - 52) * 8.4) / 70 = {new_sand_val}"

    with open(image_path, "rb") as f:
        img_data = f.read()

    html = f"""
    <html><body>
      <img src="cid:prescription_img" style="display:block;"><br>
      <pre style="font-family:monospace;font-size:13px;">
Last {boxes} Boxes Total CSI  :  {_fmt(results.get('csi'))}
SMR                           :  {_fmt(results.get('smr'))}
{group_display} Bentonite Pred :  {_fmt(results.get('finalPrediction'))}
{new_sand_line}
      </pre>
    </body></html>
    """

    msg = MIMEMultipart("related")
    msg["From"]    = SENDER
    msg["To"]      = RECIPIENT
    msg["Subject"] = SUBJECT

    msg.attach(MIMEText(html, "html"))

    img_mime = MIMEImage(img_data)
    img_mime.add_header("Content-ID", "<prescription_img>")
    img_mime.add_header("Content-Disposition", "inline")
    msg.attach(img_mime)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER, APP_PASSWORD)
            server.sendmail(SENDER, RECIPIENT, msg.as_string())
        print(f"Prescription email sent to {RECIPIENT}")
    except Exception as e:
        print(f"Email sending failed: {e}")


send_prescription_email(image_path, r, n, config2["group_name"], newsand)
